# -*- coding: utf-8 -*-
"""
Created on Fri Jun  5 11:59:03 2015

@author: mannion2
"""

from openpyxl import load_workbook
# Read in the file you want to open....make sure your file is in the same directory as this program
wb = load_workbook(filename = 'OpenpyxlTest.xlsx')
# Determine the particular sheet your would like to read in
sheet_ranges = wb['Sheet1']
# Read in a value
print(sheet_ranges['A2'].value)
# Read in all values in a row or column
print('Data in column A')
for c in range(3):
    cell = sheet_ranges.cell(row=c,column=0)
    print(cell.value)
    
    
    
    