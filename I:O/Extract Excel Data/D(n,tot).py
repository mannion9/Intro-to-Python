from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import numpy as np

# Read in 
wb = load_workbook(filename = 'D_CrossSection_Total.xlsx')
# Determine the particular sheet your would like to read in
sheet_ranges = wb['Sheet1']
# Read in all values in a row or column
Energies = []
XC = []
for i in range(13,447):
    energy , cross_section = sheet_ranges.cell(row=i,column=1) , sheet_ranges.cell(row=i,column=2)
    #cross_section = sheet_ranges.cell(row=c,column=2)    
    Energies.append(energy.value)
    XC.append(cross_section.value)

# Plot Raw Data
plt.figure(1)
plt.scatter(Energies,XC)
plt.xscale('log'),plt.yscale('log')
plt.axis([min(Energies),max(Energies),min(XC),max(XC)])
plt.xlabel('Energy (MeV)')
plt.ylabel('Cross Section (b)')
plt.title('Cross Section for D(n,tot)')

